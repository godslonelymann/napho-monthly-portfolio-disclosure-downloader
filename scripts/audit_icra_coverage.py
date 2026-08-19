#!/usr/bin/env python3
"""Audit downloaded monthly portfolio files against the ICRA scheme universe.

The script answers two questions for a given period (default ``2026-05``):

1.  For every file under ``data/raw/<amc>/<period>/``, which ICRA scheme(s)
    does it represent, and did we manage to match them?
2.  Which ICRA schemes for that period are not represented by any downloaded
    file at all?

Matching is deliberately conservative.  A scheme is only *matched* when a
normalised key resolves to exactly one ICRA scheme; anything that resolves to
more than one candidate, or that only survives an aggressive normalisation, is
reported as ``ambiguous`` rather than being silently accepted.

Match precedence (highest first)::

    amfi_code            exact AMFI code found in the file
    icra_name_exact      normalised scheme name == normalised ICRA Fund_Name
    icra_name_reduced    same, after dropping bracketed descriptors
    scheme_master_alias  normalised name == a ``Scheme Master`` scheme name
    amc_prefix_alias     same, after normalising the AMC prefix
    plan_option_reduced  same, after dropping trailing plan/option terms
    amfi_navall_code     name resolved to an AMFI code via NAVAll.txt

Usage::

    python scripts/audit_icra_coverage.py
    python scripts/audit_icra_coverage.py --period 2026-05 --amc axis --verbose
    python scripts/audit_icra_coverage.py --refresh-cache --no-amfi
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Iterator, Sequence

ROOT = Path(__file__).resolve().parents[1]

DEFAULT_PERIOD = "2026-05"
DEFAULT_ICRA = ROOT / "ICRA_Sample.xlsx"
DEFAULT_RAW = ROOT / "data" / "raw"
DEFAULT_OUT = ROOT / "outputs"
DEFAULT_CACHE = ROOT / "outputs" / ".cache"

PORTFOLIO_SHEET_TEMPLATE = "Portfolio Data_{month}"
SCHEME_MASTER_SHEET = "Scheme Master"

AMFI_NAVALL_URL = "https://portal.amfiindia.com/spages/NAVAll.txt"

# ``--`` is the workbook's sentinel for "no AMFI code assigned".
NULL_CODE_TOKENS = {"", "--", "-", "n.a.", "na", "nan", "none", "null"}


# ---------------------------------------------------------------------------
# Normalisation
# ---------------------------------------------------------------------------

_DASHES = dict.fromkeys(map(ord, "‐‑‒–—―−"), "-")
_QUOTES = dict.fromkeys(map(ord, "‘’ʼ‛"), "'")
_DQUOTES = dict.fromkeys(map(ord, "“”"), '"')

# Descriptive tails that AMC workbooks append to a scheme title.  Everything
# from the first match onwards is dropped.
_TAIL_MARKERS = re.compile(
    r"(?i)("
    r"\b(an?|the)\s+(open|close|closed)[\s-]*ended\b"
    r"|\b(open|close|closed)[\s-]*ended\s+(scheme|fund|equity|debt|hybrid|index|solution)"
    r"|\bmonthly\s+portfolio\b"
    r"|\bportfolio\s+(statement|disclosure|as\s+on|as\s+at)\b"
    r"|\bas\s+(on|at)\b"
    r"|\bfor\s+the\s+period\b"
    r"|\bfor\s+the\s+month\b"
    r"|\bfor\s+(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s*\d{2,4}\b"
    r"|\bprovisional\s+and\s+unaudited\b"
    r"|\bhalf[\s-]*yearly\s+portfolio\b"
    r")"
)

# Leading labels a cell may carry before the actual scheme name.  Two shapes
# occur: a "Label: name" prefix, and the "MONTHLY PORTFOLIO STATEMENT OF
# <name>" banner most AMCs print above the holdings table.
_LEAD_LABELS = re.compile(
    r"(?i)^\s*(?:"
    r"(?:scheme\s*name|scheme|fund\s*name|fund|name\s+of\s+the\s+scheme"
    r"|name\s+of\s+scheme|portfolio\s+of|monthly\s+portfolio\s+of|portfolio)"
    r"\s*[:\-\u2013]\s*"
    r"|(?:provisional\s+and\s+unaudited\s+)?(?:monthly\s+|half[\s-]*yearly\s+)?"
    r"portfolio\s+(?:statement|disclosure|details)?\s*(?:of|for)\s+(?:the\s+)?"
    r"(?:scheme\s+)?"
    r")"
)

# "WCAR-THE WEALTH COMPANY ARBITRAGE FUND" - an internal scheme code glued to
# the front of the name.
_CODE_PREFIX = re.compile(r"^[A-Z0-9][A-Z0-9&]{1,9}\s*[-:]\s*(?=[A-Za-z])")

_SCHEME_MARKER = re.compile(r"(?i)^\s*scheme\s*(name)?\s*[:\-]")
# Some infrastructure-debt workbooks use a bare series label as the boundary
# between stacked scheme tables rather than spelling out ``SCHEME:``.
_IMPLICIT_SCHEME_MARKER = re.compile(
    r"(?i)^\s*(?:IL\s*&\s*FS\s+)?IDF\s+Series\s+\d+\s*[- ]?\s*[A-Z]\s*$"
)

# Trailing plan / option / frequency terms, including the abbreviations that
# show up in the ICRA ``Scheme Master`` sheet.
_PLAN_TERMS = (
    r"direct|regular|institutional|instl|retail|super\s+institutional|discontinued"
    r"|unclaimed|segregated"
)
_OPTION_TERMS = (
    r"growth|idcw|income\s+distribution\s+cum\s+capital\s+withdrawal|dividend|div"
    r"|bonus|payout|pay\s*out|reinvestment|re\s*investment|reinv|appreciation"
    r"|dap|drp|adjusted\s*nav|cumulative"
)
_FREQ_TERMS = (
    r"daily|dly|weekly|wkly|fortnightly|fortnight|ftly|monthly|mthly|mnthly"
    r"|quarterly|qtly|qly|half\s*yearly|half\s*yly|hly|yearly|yly|annual|annually"
    r"|adhoc|ad\s*hoc"
)
_TRAILING_TERM = re.compile(
    r"\s+(plan|option|" + _PLAN_TERMS + "|" + _OPTION_TERMS + "|" + _FREQ_TERMS + r")$"
)

# Weakest reduction of all: some AMCs simply omit the trailing noun.
_GENERIC_TAIL = re.compile(r"\s+(fund|scheme)$")


def suffix_reductions(key: str) -> list[str]:
    """Drop a trailing generic noun ("... Infrastructure Fund" -> "... Infrastructure")."""
    out = []
    cur = key
    for _ in range(2):
        nxt = _GENERIC_TAIL.sub("", cur).strip()
        if nxt == cur or len(nxt) < 6:
            break
        cur = nxt
        out.append(cur)
    return out

_FUND_KEYWORD = re.compile(
    r"(?i)\b(fund|etf|fof|scheme|plan|index|yojana|yojna|elss|tax\s*saver|savings"
    r"|bees|series|advantage|nivesh|sip)\b"
)

# Month names are canonicalised to their three-letter form so that
# "SDL Sep 2027" and "SDL September 2027" produce the same key.
_MONTH_CANON = {}
for _abbr, _full in (
    ("jan", "january"), ("feb", "february"), ("mar", "march"), ("apr", "april"),
    ("may", "may"), ("jun", "june"), ("jul", "july"), ("aug", "august"),
    ("sep", "september"), ("oct", "october"), ("nov", "november"), ("dec", "december"),
):
    _MONTH_CANON[_abbr] = _abbr
    _MONTH_CANON[_full] = _abbr
_MONTH_CANON["sept"] = "sep"

_MONTH_WORD = re.compile(
    r"(?i)\b(" + "|".join(sorted(_MONTH_CANON, key=len, reverse=True)) + r")\b"
)

# Table furniture that is never a scheme name.  Used only when choosing which
# raw string to show as "detected scheme" for a unit we could not resolve.
_SECTION_NOISE = re.compile(
    r"(?i)^(\(?[a-z]\)?\s+)?("
    r"listed|unlisted|privately\s+placed|awaiting\s+listing|sub\s*total|total"
    r"|grand\s+total|equity|debt|money\s+market|government|treasury|corporate"
    r"|net\s|cash|others?|margin|reverse\s+repo|triparty|clearing|units?\s+of"
    r"|derivat|hedg|name\s+of|company|instrument|isin|rating|quantity|market"
    r"|yield|maturity|coupon|industry|portfolio|scheme\s+riskometer|back\s+to"
    r"|aggregate|notes?|disclaimer|mutual\s+fund\s+investments"
    r")\b"
)

_ISIN = re.compile(r"^IN[EFD][0-9A-Z]{9}$")
_AMFI_CODE = re.compile(r"^\d{5,6}$")


def _pre(text: str) -> str:
    """Shared unicode/punctuation clean-up applied before every other step."""
    s = unicodedata.normalize("NFKD", str(text))
    s = s.translate(_DASHES).translate(_QUOTES).translate(_DQUOTES)
    s = s.replace(" ", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def canonical(text: str) -> str:
    """Collapse a scheme name to its comparison key.

    Case, punctuation and whitespace are removed, ``&`` is spelled out, and any
    trailing marketing/date descriptor is cut off.  Apostrophes are deleted
    rather than spaced (``Axis Children's Fund`` -> ``axis childrens fund``) and
    letter/digit boundaries are split so ``IBX50:50`` and ``IBX 50:50`` agree.
    """
    s = _pre(text)
    s = _LEAD_LABELS.sub("", s)
    s = _TAIL_MARKERS.split(s)[0]
    s = s.replace("&", " and ")
    s = s.replace("+", " plus ")
    s = s.replace("'", "")
    s = re.sub(r"[^0-9A-Za-z]+", " ", s)
    s = re.sub(r"(?<=[A-Za-z])(?=\d)|(?<=\d)(?=[A-Za-z])", " ", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    s = _FOF.sub("fof", s)
    s = _FMP.sub("fixed maturity plan", s)
    s = " ".join(_MONTH_CANON.get(tok, tok) for tok in s.split())
    s = _roman_to_arabic(s)
    return re.sub(r"\s+", " ", s).strip()


# "Fund of Fund(s)" and "FoF" are the same thing; AMCs use both spellings.
_FOF = re.compile(r"\bfund of funds?\b")
# The industry-standard abbreviation; ICRA and the AMCs disagree on which to use.
_FMP = re.compile(r"\bfmp\b")

_ROMAN = re.compile(r"^m{0,3}(cm|cd|d?c{0,3})(xc|xl|l?x{0,3})(ix|iv|v?i{0,3})$")
_ROMAN_VALUES = {"i": 1, "v": 5, "x": 10, "l": 50, "c": 100, "d": 500, "m": 1000}
# Only converted directly after one of these words, so that a stray "MIX" or a
# series letter such as "- C" is never mistaken for a numeral.
_ROMAN_TRIGGERS = {"series", "plan", "fmp", "no"}


def _roman_to_arabic(key: str) -> str:
    """Rewrite "Series XLIII" as "series 43" so numbering styles agree."""
    tokens = key.split()
    out: list[str] = []
    for i, tok in enumerate(tokens):
        prev = tokens[i - 1] if i else ""
        if (
            len(tok) >= 2
            and prev in _ROMAN_TRIGGERS
            and _ROMAN.match(tok)
        ):
            total = 0
            prev_val = 0
            for ch in reversed(tok):
                val = _ROMAN_VALUES[ch]
                total += -val if val < prev_val else val
                prev_val = max(prev_val, val)
            out.append(str(total))
        else:
            out.append(tok)
    return " ".join(out)


def drop_brackets(text: str) -> str:
    """Remove bracketed descriptors, e.g. ``Fund (An open ended scheme)``."""
    s = _pre(text)
    prev = None
    while prev != s:
        prev = s
        s = re.sub(r"\([^()]*\)", " ", s)
        s = re.sub(r"\[[^\[\]]*\]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


_TENOR_TAIL = re.compile(
    r"\s+\d{1,5}\s*(days?|d|months?|mths?|m|years?|yrs?|y)$"
)


def strip_tenor(key: str) -> str:
    """Drop a trailing tenor such as "- 3668 Days" from a fixed-maturity name."""
    prev = None
    cur = key
    while prev != cur:
        prev = cur
        cur = _TENOR_TAIL.sub("", cur).strip()
    return cur


def plan_reductions(key: str) -> list[str]:
    """Progressively strip trailing plan/option/frequency terms.

    Returns the chain longest-first so callers can prefer the least-reduced
    form.  ``Aditya Birla Sun Life Medium Term Plan`` therefore still matches on
    its full key before the ``plan`` suffix is considered noise.
    """
    out: list[str] = []
    cur = key
    for _ in range(8):
        nxt = _TRAILING_TERM.sub("", cur).strip()
        if nxt == cur or not nxt:
            break
        cur = nxt
        out.append(cur)
    return out


def is_specific(key: str) -> str:
    """Reject keys too generic to identify a scheme.

    Reductions can erode a name down to a single common word - stripping the
    AMC prefix and the trailing "Fund" turns "HDFC Income Fund" into "income",
    which then matches any stray "Income" header cell.  Two tokens is the floor.
    """
    return len(key) >= 10 and len(key.split()) >= 2


def token_set(key: str) -> str:
    """Order-independent form of a key: its tokens, sorted.

    "HDFC FMP 1861D March 2022" and "HDFC Fixed Maturity Plan - March 2022 -
    1861D" carry identical tokens in a different order.  This is still exact
    equality, not similarity - a single differing token (1861D vs 1876D) makes
    the keys unequal.
    """
    return " ".join(sorted(key.split()))


def despace(key: str) -> str:
    """Whitespace-insensitive form of a key.

    AMCs are inconsistent about compound words ("LargeMidcap" vs
    "Large Midcap", "Flexicap" vs "Flexi Cap").  Collapsing the spaces is still
    a deterministic normalisation, and any two ICRA schemes that collide under
    it end up in the same index bucket and are therefore reported as ambiguous
    rather than matched.
    """
    return key.replace(" ", "")


@dataclass(frozen=True)
class Candidate:
    """A normalised key derived from some raw text, with its provenance.

    ``order`` ranks provenance: manifest metadata first, then sheet titles in
    row order, then the sheet name, then the filename.  Portfolio workbooks
    sometimes mention an unrelated scheme in a footer, so when two sources
    disagree the earlier one wins instead of the pair being called ambiguous.
    """

    key: str
    tier: str
    raw: str
    source: str
    order: int = 1000


def build_candidates(
    raw: str, source: str, require_keyword: bool = True, order: int = 1000
) -> list[Candidate]:
    """Turn one raw string into the ordered set of keys worth looking up.

    ``require_keyword`` is a cheap guard against turning every holdings line
    into a lookup.  It is relaxed for cells in a sheet's header area, where the
    volume is small and a scheme title may legitimately omit the word "Fund"
    (e.g. "Navi Nifty SmallCap250 Momentum Qlty 100").
    """
    raw = _pre(raw)
    if not raw or len(raw) > 600:
        return []
    if require_keyword and not _FUND_KEYWORD.search(raw):
        return []

    out: list[Candidate] = []
    seen: set[str] = set()

    if not re.search(r"[A-Za-z]", raw):
        return []

    def add(key: str, tier: str) -> None:
        # The cap is applied to the trimmed key so that a long banner such as
        # "MONTHLY PORTFOLIO STATEMENT OF <name> AS ON ... (An open ended ...)"
        # is still usable once the boilerplate has been cut away.
        if key and 5 <= len(key) <= 120 and key not in seen:
            seen.add(key)
            out.append(Candidate(key, tier, raw, source, order))

    exact = canonical(raw)
    add(exact, "exact")

    reduced = canonical(drop_brackets(raw))
    add(reduced, "reduced")

    decoded = _CODE_PREFIX.sub("", raw)
    if decoded != raw:
        add(canonical(decoded), "exact")
        add(canonical(drop_brackets(decoded)), "reduced")

    for base in list(seen):
        tenorless = strip_tenor(base)
        if tenorless != base:
            add(tenorless, "reduced")

    for base in list(seen):
        for red in plan_reductions(base):
            add(red, "plan")

    for base in list(seen):
        for red in suffix_reductions(base):
            add(red, "suffix")
    return out


def amc_prefix_variants(key: str, prefixes: Sequence[str]) -> list[str]:
    """Strip any known AMC prefix from ``key`` (longest prefix first)."""
    out = []
    for pref in prefixes:
        if key.startswith(pref + " ") and len(key) > len(pref) + 1:
            out.append(key[len(pref) + 1 :].strip())
    return out


# ---------------------------------------------------------------------------
# ICRA reference data
# ---------------------------------------------------------------------------


@dataclass
class IcraScheme:
    scheme_id: int
    amfi_code: str  # "" when ICRA has no code
    fund_name: str
    mf_name: str
    aliases: list[str] = field(default_factory=list)


@dataclass
class IcraUniverse:
    schemes: list[IcraScheme]
    by_code: dict[str, int]
    # (mf_name, tier, key) -> scheme ids ; mf_name "" is the global scope
    index: dict[tuple[str, str, str], set[int]]
    mf_names: list[str]
    prefixes: dict[str, list[str]]

    def lookup(self, mf_name: str, tier: str, key: str) -> set[int]:
        return self.index.get((mf_name, tier, key), set())


def _clean_code(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    s = str(value).strip()
    if s.lower() in NULL_CODE_TOKENS:
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _iter_xlsx_rows(path: Path, sheet: str) -> Iterator[tuple]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        yield from ws.iter_rows(values_only=True)
    finally:
        wb.close()


def load_icra(
    icra_path: Path, period: str, cache_dir: Path, refresh: bool
) -> tuple[list[dict], list[dict]]:
    """Read the portfolio + scheme-master sheets, caching the extract as JSON."""
    month_label = _period_month_label(period)
    sheet = PORTFOLIO_SHEET_TEMPLATE.format(month=month_label)
    stat = icra_path.stat()
    cache = cache_dir / f"icra_{period}_{int(stat.st_mtime)}_{stat.st_size}.json"

    if cache.exists() and not refresh:
        data = json.loads(cache.read_text())
        return data["portfolio"], data["master"]

    portfolio: dict[tuple[str, str], dict] = {}
    header_seen = False
    for row in _iter_xlsx_rows(icra_path, sheet):
        if not header_seen:
            header_seen = True
            continue
        if row is None or len(row) < 13:
            continue
        code = _clean_code(row[0])
        name = row[12]
        if not name:
            continue
        name = _pre(name)
        key = (code, name)
        if key not in portfolio:
            portfolio[key] = {"amfi_code": code, "fund_name": name, "rows": 0}
        portfolio[key]["rows"] += 1

    master: list[dict] = []
    header_seen = False
    for row in _iter_xlsx_rows(icra_path, SCHEME_MASTER_SHEET):
        if not header_seen:
            header_seen = True
            continue
        if row is None or len(row) < 3 or not row[1]:
            continue
        master.append(
            {
                "amfi_code": _clean_code(row[0]),
                "scheme_name": _pre(row[1]),
                "mf_name": _pre(row[2]) if row[2] else "",
            }
        )

    cache_dir.mkdir(parents=True, exist_ok=True)
    for stale in cache_dir.glob(f"icra_{period}_*.json"):
        stale.unlink()
    cache.write_text(json.dumps({"portfolio": list(portfolio.values()), "master": master}))
    return list(portfolio.values()), master


def _period_month_label(period: str) -> str:
    year, month = period.split("-")
    months = [
        "Jan", "Feb", "Mar", "Apr", "May", "Jun",
        "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
    ]
    return f"{months[int(month) - 1]}{year}"


def build_universe(portfolio: list[dict], master: list[dict]) -> IcraUniverse:
    """Index the ICRA schemes by code, name, and Scheme Master alias."""
    # Scheme Master, grouped by its plan-stripped base name.
    master_by_code: dict[str, dict] = {}
    master_by_base: dict[str, list[dict]] = defaultdict(list)
    for row in master:
        if row["amfi_code"]:
            master_by_code.setdefault(row["amfi_code"], row)
        base = canonical(row["scheme_name"])
        chain = [base] + plan_reductions(base)
        for key in chain:
            master_by_base[key].append(row)

    schemes: list[IcraScheme] = []
    for idx, row in enumerate(sorted(portfolio, key=lambda r: (r["fund_name"], r["amfi_code"]))):
        base = canonical(row["fund_name"])
        mf_name = ""
        if row["amfi_code"] and row["amfi_code"] in master_by_code:
            mf_name = master_by_code[row["amfi_code"]]["mf_name"]
        if not mf_name:
            for cand in master_by_base.get(base, []):
                if cand["mf_name"]:
                    mf_name = cand["mf_name"]
                    break
        aliases = sorted(
            {
                m["scheme_name"]
                for m in master_by_base.get(base, [])
                if not mf_name or m["mf_name"] == mf_name
            }
        )
        if row["amfi_code"]:
            aliases.extend(
                m["scheme_name"]
                for m in master
                if m["amfi_code"] == row["amfi_code"]
            )
        schemes.append(
            IcraScheme(
                scheme_id=idx,
                amfi_code=row["amfi_code"],
                fund_name=row["fund_name"],
                mf_name=mf_name,
                aliases=sorted(set(aliases)),
            )
        )

    by_code: dict[str, int] = {}
    index: dict[tuple[str, str, str], set[int]] = defaultdict(set)
    mf_names = sorted({s.mf_name for s in schemes if s.mf_name})
    prefixes = {mf: sorted(amc_prefix_keys(mf), key=len, reverse=True) for mf in mf_names}

    def register(scheme: IcraScheme, tier: str, key: str) -> None:
        if not is_specific(key):
            return
        for form in {key, despace(key)}:
            index[(scheme.mf_name, tier, form)].add(scheme.scheme_id)
            index[("", tier, form)].add(scheme.scheme_id)
        for pref in prefixes.get(scheme.mf_name, []):
            for shortened in amc_prefix_variants(key, [pref]):
                for form in {shortened, despace(shortened)}:
                    index[(scheme.mf_name, "amc_prefix", form)].add(scheme.scheme_id)

    for scheme in schemes:
        if scheme.amfi_code:
            by_code.setdefault(scheme.amfi_code, scheme.scheme_id)
        base = canonical(scheme.fund_name)
        register(scheme, "exact", base)
        reduced = canonical(drop_brackets(scheme.fund_name))
        if reduced != base:
            register(scheme, "reduced", reduced)
        for form in {base, reduced}:
            tenorless = strip_tenor(form)
            if tenorless != form:
                register(scheme, "reduced", tenorless)
            register(scheme, "tokenset", token_set(form))
        for red in plan_reductions(base):
            register(scheme, "plan", red)
        for red in suffix_reductions(base):
            register(scheme, "suffix", red)
        for alias in scheme.aliases:
            abase = canonical(alias)
            register(scheme, "alias", abase)
            areduced = canonical(drop_brackets(alias))
            if areduced != abase:
                register(scheme, "alias", areduced)
            for red in plan_reductions(abase):
                register(scheme, "alias", red)

    # Scheme Master codes that resolve to a portfolio scheme widen code lookup.
    base_to_scheme: dict[str, set[int]] = defaultdict(set)
    for scheme in schemes:
        base_to_scheme[canonical(scheme.fund_name)].add(scheme.scheme_id)
    for row in master:
        if not row["amfi_code"] or row["amfi_code"] in by_code:
            continue
        base = canonical(row["scheme_name"])
        hits: set[int] = set()
        for key in [base] + plan_reductions(base):
            hits |= base_to_scheme.get(key, set())
            if hits:
                break
        if len(hits) == 1:
            by_code[row["amfi_code"]] = next(iter(hits))

    return IcraUniverse(schemes, by_code, dict(index), mf_names, prefixes)


def amc_prefix_keys(mf_name: str) -> set[str]:
    """Normalised prefixes an AMC may use in front of its scheme names."""
    base = canonical(mf_name)
    out = {base}
    for suffix in (" mutual fund", " mf"):
        if base.endswith(suffix):
            out.add(base[: -len(suffix)].strip())
    stem = next(iter(sorted(out, key=len)))
    extra = MANUAL_PREFIXES.get(mf_name, [])
    out.update(canonical(x) for x in extra)
    out.add(stem + " mf")
    return {p for p in out if len(p) >= 3}


# Prefix spellings AMCs use in their own files that ICRA does not.
MANUAL_PREFIXES: dict[str, list[str]] = {
    "Aditya Birla Sun Life Mutual Fund": ["ABSL", "Birla Sun Life", "Aditya Birla"],
    "Baroda BNP Paribas Mutual Fund": ["Baroda BNP", "BNP Paribas", "BOBBNP", "BBP"],
    "Bank of India Mutual Fund": ["BOI", "BOI AXA"],
    "Canara Robeco Mutual Fund": ["Canara"],
    "DSP Mutual Fund": ["DSP BlackRock"],
    "Franklin Templeton Mutual Fund": ["Franklin India", "Franklin", "Templeton India"],
    "HSBC Mutual Fund": ["HSBC India"],
    "ICICI Prudential Mutual Fund": ["ICICI Pru", "ICICI"],
    "IL & FS Mutual Fund": ["IL&FS", "ILFS"],
    "JM Financial Mutual Fund": ["JM"],
    "Jio BlackRock Mutual Fund": ["JioBlackRock", "Jio Black Rock"],
    "Kotak Mahindra Mutual Fund": ["Kotak"],
    "LIC Mutual Fund": ["LIC MF", "LICMF"],
    "Mahindra Manulife Mutual Fund": ["Mahindra"],
    "Mirae Asset Mutual Fund": ["Mirae"],
    "Motilal Oswal Mutual Fund": ["Motilal"],
    "Nippon India Mutual Fund": ["Nippon", "Reliance"],
    "PGIM India Mutual Fund": ["PGIM"],
    "PPFAS Mutual Fund": ["Parag Parikh", "PPFAS"],
    "SBI Mutual Fund": ["SBI Magnum"],
    "The Wealth Company Mutual Fund": ["Wealth Company", "The Wealth Company"],
    "Trust Mutual Fund": ["TRUSTMF", "Trust MF"],
    "UTI Mutual Fund": ["UTI"],
    "WhiteOak Capital Mutual Fund": ["WhiteOak", "WOC", "White Oak"],
    "360 ONE Mutual Fund": ["360 ONE", "IIFL"],
    "Quantum Mutual Fund": ["Quantum"],
    "Angel One Mutual Fund": ["Angel One"],
}

# raw directory name -> ICRA "MF Name"; only the cases that do not fall out of
# a straightforward slug comparison.
MANUAL_AMC_MAP = {
    "wealth_company": "The Wealth Company Mutual Fund",
    # slug() strips all non-alphanumerics, so "bandhan/monthly-half-yearly"
    # collapses to "bandhanmonthlyhalfyearly" -- nothing close enough to
    # "bandhanmutualfund" for the fuzzy startswith match below to accept,
    # hence explicit entries here. Named after each source page's own URL
    # path segment (see verified/Bandhan_Mutual_Fund.py): the per-scheme
    # disclosure page is ".../scheme-portfolios/monthly-half-yearly", the
    # consolidated summary page is ".../portfolio-summary/monthly".
    "bandhan/monthly-half-yearly": "Bandhan Mutual Fund",
    "bandhan/monthly": "Bandhan Mutual Fund",
}


def map_amc_dirs(dirs: Iterable[str], mf_names: Sequence[str]) -> dict[str, str]:
    """Resolve ``data/raw`` directory names to ICRA ``MF Name`` values."""

    def slug(text: str) -> str:
        return re.sub(r"[^a-z0-9]", "", text.lower().replace("mutual fund", ""))

    lookup = {slug(m): m for m in mf_names}
    out: dict[str, str] = {}
    for d in dirs:
        if d in MANUAL_AMC_MAP:
            out[d] = MANUAL_AMC_MAP[d]
            continue
        key = slug(d)
        if key in lookup:
            out[d] = lookup[key]
            continue
        hits = [m for k, m in lookup.items() if k.startswith(key) or key.startswith(k)]
        out[d] = hits[0] if len(hits) == 1 else ""
    return out


# ---------------------------------------------------------------------------
# AMFI NAVAll.txt (tier-4 fallback)
# ---------------------------------------------------------------------------


def load_amfi(cache_dir: Path, refresh: bool, enabled: bool) -> dict[str, set[str]]:
    """Return ``canonical(scheme name) -> AMFI codes`` from NAVAll.txt.

    Names are indexed both verbatim and with their plan/option suffix removed.
    A base scheme normally has several AMFI codes (one per plan/option), so the
    caller translates every code to an ICRA scheme and only accepts the result
    when *all mapped codes* converge on one scheme.  This keeps the fallback
    deterministic without discarding the common multi-code case.
    """
    if not enabled:
        return {}
    cache = cache_dir / "amfi_navall.txt"
    text = ""
    if cache.exists() and not refresh:
        text = cache.read_text(encoding="utf-8", errors="replace")
    else:
        try:
            import urllib.request

            req = urllib.request.Request(
                AMFI_NAVALL_URL, headers={"User-Agent": "portfolio-audit/1.0"}
            )
            with urllib.request.urlopen(req, timeout=90) as resp:
                text = resp.read().decode("utf-8", errors="replace")
            cache_dir.mkdir(parents=True, exist_ok=True)
            cache.write_text(text, encoding="utf-8")
        except Exception as exc:  # network is optional
            print(f"  ! AMFI NAVAll.txt unavailable ({exc}); skipping tier 4", file=sys.stderr)
            if cache.exists():
                text = cache.read_text(encoding="utf-8", errors="replace")
            else:
                return {}

    by_key: dict[str, set[str]] = defaultdict(set)
    for line in text.splitlines():
        if line.count(";") < 5:
            continue
        parts = line.split(";")
        code = _clean_code(parts[0])
        name = _pre(parts[3])
        if not code or not _AMFI_CODE.match(code) or not name:
            continue
        base = canonical(name)
        for key in [base] + plan_reductions(base):
            by_key[key].add(code)
    return dict(by_key)


# ---------------------------------------------------------------------------
# Workbook reading
# ---------------------------------------------------------------------------

# Sheets that are never a scheme portfolio.  "Derivative" sheets in particular
# are a cross-scheme annexe, so counting them as units would double-count.
_SKIP_SHEET = re.compile(
    r"(?i)^\s*(index|contents?|notes?|disclaimer|disclosure|cover|legend|glossary"
    r"|risk[\s-]*o[\s-]*meter.*|dividend\s+history|.*performance|.*risk-?o-?meter"
    r"|expense\s+ratio|scheme'?s?\s+aum|investment\s+objective|abbreviation.*"
    r"|derivatives?.*|.*\s+notes|common\s+notes.*|notes\s+to\s+.*"
    r"|.*derivative\s+disclosure.*)\s*$"
)
_DATA_HEADER = re.compile(
    r"(?i)(name\s+of\s+the\s+instrument|company/?issuer|instrument\s+name|isin"
    r"|%\s*to\s*(net\s*)?(asset|nav)|market\s*/?\s*fair\s*value|net\s+assets)"
)


@dataclass
class SheetScan:
    name: str
    titles: list[tuple[int, str]]           # (row index, text)
    markers: list[tuple[int, str]]          # rows introduced by "SCHEME:" etc.
    has_isin: bool = False
    has_header: bool = False
    rows: int = 0
    top_rows: int = 15                      # size of the relaxed header region
    pairs: list[list[str]] = field(default_factory=list)  # short index-sheet rows

    @property
    def is_data_sheet(self) -> bool:
        return self.has_isin or self.has_header


def _row_strings(values: Sequence) -> tuple[list[str], bool]:
    out: list[str] = []
    has_isin = False
    for v in values:
        if v is None:
            continue
        if isinstance(v, str):
            s = v.strip()
        elif isinstance(v, (int, float)):
            s = str(v)
        else:
            s = str(v).strip()
        if not s:
            continue
        out.append(s)
        if not has_isin and len(s) == 12 and _ISIN.match(s.upper()):
            has_isin = True
    return out, has_isin


def scan_sheet(name: str, rows: Iterator[Sequence], top_rows: int = 15) -> SheetScan:
    """Collect title-like strings from a sheet without loading it all in memory.

    A cell is a title candidate when its row carries no ISIN (so it is not a
    holdings line) and either sits in the sheet header area or is one of at
    most four populated cells on the row.
    """
    scan = SheetScan(name=name, titles=[], markers=[], top_rows=top_rows)
    for idx, raw in enumerate(rows):
        strings, has_isin = _row_strings(raw)
        if has_isin:
            scan.has_isin = True
        if not strings:
            continue
        scan.rows = idx + 1
        if not scan.has_header and len(strings) >= 3:
            joined = " | ".join(strings[:12])
            if _DATA_HEADER.search(joined):
                scan.has_header = True
        if has_isin:
            continue
        if 2 <= len(strings) <= 4 and len(scan.pairs) < 400:
            scan.pairs.append(strings)
        if idx >= top_rows and len(strings) > 4:
            continue
        for cell in strings:
            # The upper bound is generous: several AMCs wrap the scheme name in
            # a long banner that build_candidates() trims down later.
            if len(cell) < 6 or len(cell) > 600:
                continue
            if _SCHEME_MARKER.match(cell) or _IMPLICIT_SCHEME_MARKER.match(cell):
                scan.markers.append((idx, cell))
                scan.titles.append((idx, cell))
            elif idx < top_rows or _FUND_KEYWORD.search(cell):
                scan.titles.append((idx, cell))
    return scan


def read_workbook(path: Path) -> tuple[str, list[SheetScan], str]:
    """Return ``(kind, sheet scans, error)`` for one downloaded file."""
    try:
        data = path.read_bytes()
    except OSError as exc:
        return "", [], f"unreadable: {exc}"

    if data[:2] == b"PK":
        try:
            import openpyxl

            wb = openpyxl.load_workbook(
                io.BytesIO(data), read_only=True, data_only=True
            )
        except Exception as exc:
            return "xlsx", [], f"openpyxl failed: {type(exc).__name__}: {exc}"
        scans = []
        try:
            for ws in wb.worksheets:
                try:
                    scans.append(scan_sheet(ws.title, ws.iter_rows(values_only=True)))
                except Exception as exc:
                    scans.append(SheetScan(name=ws.title, titles=[], markers=[]))
                    scans[-1].rows = -1
                    _ = exc
        finally:
            wb.close()
        return "xlsx", scans, ""

    if data[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        try:
            import xlrd

            wb = xlrd.open_workbook(file_contents=data)
        except ModuleNotFoundError:
            # The bundled runtime intentionally does not ship xlrd.  Convert
            # the legacy workbook in an isolated temporary directory when a
            # local LibreOffice binary is available.  The downloaded source is
            # never touched.
            office = shutil.which("soffice") or shutil.which("libreoffice")
            if not office:
                return "xls", [], "legacy .xls reader unavailable (xlrd/LibreOffice not found)"
            try:
                with tempfile.TemporaryDirectory(prefix="portfolio-audit-xls-") as tmp:
                    proc = subprocess.run(
                        [office, "--headless", "--convert-to", "xlsx", "--outdir", tmp, str(path)],
                        check=False,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.STDOUT,
                        text=True,
                        timeout=120,
                    )
                    converted = Path(tmp) / f"{path.stem}.xlsx"
                    if proc.returncode or not converted.exists():
                        detail = re.sub(r"\s+", " ", proc.stdout).strip()[:300]
                        return "xls", [], f"LibreOffice conversion failed: {detail}"
                    _, scans, error = read_workbook(converted)
                    return "xls", scans, error
            except Exception as exc:
                return "xls", [], f"LibreOffice conversion failed: {type(exc).__name__}: {exc}"
        except Exception as exc:
            return "xls", [], f"xlrd failed: {type(exc).__name__}: {exc}"
        scans = []
        for sh in wb.sheets():
            scans.append(
                scan_sheet(sh.name, (sh.row_values(i) for i in range(sh.nrows)))
            )
        return "xls", scans, ""

    head = data[:512].lstrip().lower()
    if head.startswith(b"<") or b"<html" in head:
        return "html", [], "unreadable: HTML content, not a workbook"
    if head.startswith(b"%pdf"):
        return "pdf", [], "unreadable: PDF, not a workbook"
    return "unknown", [], f"unreadable: unrecognised signature {data[:8]!r}"


# ---------------------------------------------------------------------------
# Scheme units and matching
# ---------------------------------------------------------------------------

METHOD_CONFIDENCE = {
    "amfi_code": 1.00,
    "icra_name_exact": 0.98,
    "icra_name_reduced": 0.93,
    "token_set_equal": 0.91,
    "scheme_master_alias": 0.90,
    "amc_prefix_alias": 0.86,
    "plan_option_reduced": 0.80,
    "generic_suffix_reduced": 0.78,
    "amfi_navall_code": 0.75,
}

# Both sides of a comparison carry a normalisation tier.  The strength of a hit
# is the weaker of the two, so an exact ICRA name reached through an
# aggressively reduced candidate is still reported as a reduced match.
TIER_RANK = {
    "exact": 6, "reduced": 5, "tokenset": 4, "alias": 3,
    "amc_prefix": 2, "plan": 1, "suffix": 0,
}
RANK_METHOD = {
    6: "icra_name_exact",
    5: "icra_name_reduced",
    4: "token_set_equal",
    3: "scheme_master_alias",
    2: "amc_prefix_alias",
    1: "plan_option_reduced",
    0: "generic_suffix_reduced",
}
INDEX_TIERS = ("exact", "reduced", "tokenset", "alias", "amc_prefix", "plan", "suffix")


# Provenance ranking used by Candidate.order (lower wins).
ORDER_MANIFEST = 0
ORDER_SHEET_TITLE = 100     # + row index
ORDER_SHEET_NAME = 1000
ORDER_FILENAME = 2000


@dataclass
class Unit:
    """One scheme-sized slice of a downloaded file."""

    label: str                      # sheet name, or sheet + marker text
    candidates: list[Candidate]
    codes: list[str]
    detected_name: str


_FILE_NOISE = re.compile(
    r"(?i)\b(monthly|month\s*end|portfolio|portfolios|disclosure|disclosures|statement"
    r"|report|isin|as\s+on|as\s+at|final|holdings?|mf|scheme|schemes|details)\b"
)
_MONTH_TOKEN = _MONTH_WORD


def filename_candidates(path: Path) -> list[Candidate]:
    """Derive keys from the filename, from least to most aggressive cleaning.

    Numbers are preserved: ``Nifty 50`` and ``Series 129`` are part of scheme
    names.  Only date-shaped and hash-shaped tokens are removed, and each stage
    is offered as its own candidate so a match can be traced back to it.
    """
    stem = re.sub(r"[_\-]+", " ", path.stem)
    stem = re.sub(r"\b[0-9a-f]{8,}\b", " ", stem)          # content hashes
    stem = re.sub(r"\b\d{1,2}[./]\d{1,2}[./]\d{2,4}\b", " ", stem)  # 31.05.2026
    stem = re.sub(r"\b\d{6,}\b", " ", stem)                # timestamps / ids
    v1 = re.sub(r"\s+", " ", stem).strip()

    v2 = re.sub(r"\s+", " ", _FILE_NOISE.sub(" ", v1)).strip()

    v3 = _MONTH_TOKEN.sub(" ", v2)
    v3 = re.sub(r"\b(19|20)\d{2}\b", " ", v3)
    v3 = re.sub(r"(?i)\b\d{1,2}(st|nd|rd|th)?\b\s*$", " ", v3)
    v3 = re.sub(r"\s+", " ", v3).strip()

    out: list[Candidate] = []
    seen: set[str] = set()
    for order, (text, src_label) in enumerate(
        ((v1, "filename"), (v2, "filename_cleaned"), (v3, "filename_dateless"))
    ):
        for cand in build_candidates(text, src_label, order=ORDER_FILENAME + order):
            if cand.key not in seen:
                seen.add(cand.key)
                out.append(cand)
    return out


def extract_codes(scan: SheetScan) -> list[str]:
    """AMFI codes explicitly labelled inside a sheet."""
    out = []
    for _, text in scan.titles:
        m = re.search(r"(?i)\bamfi\s*(scheme\s*)?code\s*[:\-]?\s*(\d{5,6})\b", text)
        if m:
            out.append(m.group(2))
    return out


def index_sheet_map(scans: list[SheetScan]) -> dict[str, str]:
    """Read a workbook's "Index" sheet as ``sheet code -> scheme name``.

    Consolidated workbooks name their tabs with internal codes ("K5",
    "AXIS113") and carry the expansion on an index tab.  Where a tab's own
    header does not spell the scheme out, this recovers it.
    """
    sheet_names = {s.name.strip().lower(): s.name for s in scans}
    out: dict[str, str] = {}
    for scan in scans:
        if not re.match(r"(?i)^\s*(index|contents?)\s*$", scan.name.strip()):
            continue
        for cells in scan.pairs:
            for i, code in enumerate(cells):
                target = sheet_names.get(code.strip().lower())
                if not target:
                    continue
                for name in cells[i + 1:] + cells[:i]:
                    name = name.strip()
                    if len(name) > 6 and not name.isdigit():
                        out.setdefault(target, name)
                        break
                break
    return out


_ASOF_DATE = re.compile(
    r"(?i)\bportfolio\s+as\s+(?:on|at)\b.*?\b"
    r"(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|"
    r"aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)"
    r"\b.*?\b((?:19|20)\d{2})\b"
)


def _sheet_is_stale(scan: SheetScan, period: str) -> bool:
    """True when a sheet explicitly labels its portfolio as another month."""
    year, month = period.split("-")
    wanted_month = int(month)
    month_numbers = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }
    for row_index, text in scan.titles:
        # Footnotes often quote a prior month's riskometer date.  Only an
        # explicit as-of label in the sheet's header can make the sheet stale.
        if row_index >= scan.top_rows:
            continue
        match = _ASOF_DATE.search(text)
        if not match:
            continue
        found_month = month_numbers[match.group(1).lower()[:3]]
        return match.group(2) != year or found_month != wanted_month
    return False


def build_units(
    path: Path, scans: list[SheetScan], manifest_scheme: str,
    period: str = DEFAULT_PERIOD,
) -> tuple[list[Unit], list[str]]:
    """Split a workbook into scheme units and list the sheets we skipped."""
    shared = filename_candidates(path)
    if manifest_scheme and manifest_scheme.lower() not in {
        "consolidated", "combined", "all funds portfolio",
    }:
        shared = build_candidates(
            manifest_scheme, "manifest", order=ORDER_MANIFEST
        ) + shared

    data_sheets = [
        s for s in scans
        if (s.is_data_sheet or len(s.markers) >= 2)
        and not _SKIP_SHEET.match(s.name)
        and not _sheet_is_stale(s, period)
    ]
    skipped = [
        s.name + (" (as-of outside audited period)" if _sheet_is_stale(s, period) else "")
        for s in scans if s not in data_sheets
    ]
    from_index = index_sheet_map(scans)

    units: list[Unit] = []
    for scan in data_sheets:
        codes = extract_codes(scan)
        sheet_cands = build_candidates(scan.name, "sheet_name", order=ORDER_SHEET_NAME)
        if scan.name in from_index:
            sheet_cands = build_candidates(
                from_index[scan.name], "index_sheet", order=ORDER_SHEET_NAME - 1
            ) + sheet_cands

        if len(scan.markers) >= 2:
            # Several schemes stacked inside one sheet (e.g. UTI's SEBI export).
            for idx, text in scan.markers:
                cands = build_candidates(
                    text, f"sheet:{scan.name}:row{idx}", order=ORDER_SHEET_TITLE + idx
                )
                units.append(
                    Unit(
                        label=f"{scan.name} @row{idx + 1}",
                        candidates=cands + sheet_cands + shared,
                        codes=codes,
                        detected_name=_LEAD_LABELS.sub("", _pre(text)) or text,
                    )
                )
            continue

        title_cands: list[Candidate] = []
        best_name = ""
        for idx, text in scan.titles:
            cands = build_candidates(
                text, f"sheet:{scan.name}:row{idx}",
                require_keyword=idx >= scan.top_rows,
                order=ORDER_SHEET_TITLE + idx,
            )
            if cands and not best_name:
                best_name = _LEAD_LABELS.sub("", _pre(text))
            title_cands.extend(cands)
        units.append(
            Unit(
                label=scan.name,
                candidates=title_cands + sheet_cands + shared,
                codes=codes,
                detected_name=best_name or scan.name,
            )
        )

    if not units:
        # No sheet looked like a portfolio table; still try filename/manifest.
        units.append(
            Unit(
                label="(file)",
                candidates=shared,
                codes=[],
                detected_name=_pre(manifest_scheme) or path.stem,
            )
        )
    return units, skipped


@dataclass
class MatchResult:
    status: str            # matched | ambiguous | unmatched
    method: str
    confidence: float
    scheme_id: int | None
    candidates: list[int]
    evidence: str
    note: str = ""
    detected: str = ""      # the raw text the decision was based on
    nearest_fund: str = ""  # closest ICRA name, for the unresolved report
    nearest_score: float = 0.0


FUZZY_REPORT_FLOOR = 0.70   # below this a near miss is not even worth listing
FUZZY_AMBIGUOUS_FLOOR = 0.86  # at/above this it is reported as an ambiguous candidate


def fuzzy_candidates(
    unit: Unit, universe: IcraUniverse, scope: str, limit: int = 3
) -> list[tuple[float, int, str]]:
    """Rank ICRA schemes by string similarity to this unit's candidate keys.

    Purely advisory: the caller never promotes these to a match, it only reports
    them so a human can adjudicate abbreviations such as "Indx" for "Index".
    """
    from difflib import SequenceMatcher

    pool = [s for s in universe.schemes if not scope or s.mf_name == scope]
    if not pool or not unit.candidates:
        return []
    # The AMC's own name is not a scheme name; scoring it would rank a house
    # banner ("ICICI Prudential Mutual Fund") against every fund it sells.
    banned = {despace(b) for b in universe.prefixes.get(scope, [])}
    banned.add(despace(canonical(scope)))
    # Advisory-only abbreviation expansion improves the candidate list without
    # changing deterministic matching.  These variants can only produce an
    # ``ambiguous`` near miss; they are never silently promoted to a match.
    def advisory_key(key: str) -> str:
        replacements = (
            (r"\bidf\b", "infrastructure debt fund"),
            (r"\bqlty\b", "quality"),
            (r"\bindx\b", "index"),
            (r"\bfin\s+servs?\b", "financial services"),
            (r"\bsec\b", "securities"),
            (r"\badvt\b", "advantage"),
            (r"\balloc\b", "allocation"),
            (r"\beq\b", "equity"),
        )
        out = key
        for pattern, replacement in replacements:
            out = re.sub(pattern, replacement, out)
        return re.sub(r"\s+", " ", out).strip()

    keys = {
        despace(advisory_key(c.key)): c.raw
        for c in unit.candidates
        if is_specific(c.key) and despace(c.key) not in banned
    }
    if not keys:
        return []

    scored: dict[int, tuple[float, str]] = {}
    for scheme in pool:
        target = despace(canonical(scheme.fund_name))
        best = 0.0
        best_raw = ""
        for key, raw in keys.items():
            ratio = SequenceMatcher(None, key, target).ratio()
            if ratio > best:
                best, best_raw = ratio, raw
        if best >= FUZZY_REPORT_FLOOR:
            scored[scheme.scheme_id] = (best, best_raw)
    ranked = sorted(
        ((v[0], sid, v[1]) for sid, v in scored.items()), key=lambda x: -x[0]
    )
    return ranked[:limit]


def match_unit(
    unit: Unit, universe: IcraUniverse, mf_name: str, amfi: dict[str, set[str]]
) -> MatchResult:
    """Resolve one unit to an ICRA scheme, or explain why we could not.

    Every candidate key is probed against every index tier; the strongest
    rank that produces hits decides the outcome.  If that rank yields more than
    one distinct ICRA scheme the unit is reported as ambiguous rather than
    guessed at - the same applies to a lone hit that belongs to another AMC.
    """
    scope = mf_name or ""

    # Tier 1 - an AMFI code printed inside the file wins outright.
    for code in unit.codes:
        sid = universe.by_code.get(code)
        if sid is not None:
            return MatchResult("matched", "amfi_code", METHOD_CONFIDENCE["amfi_code"],
                               sid, [sid], f"AMFI code {code}",
                               detected=universe.schemes[sid].fund_name)

    # Widen the candidate list with AMC-prefix-stripped forms.
    probes: list[tuple[str, str, Candidate]] = []
    prefixes = universe.prefixes.get(scope, [])
    for cand in unit.candidates:
        if not is_specific(cand.key):
            continue
        probes.append((cand.tier, cand.key, cand))
        probes.append((cand.tier, despace(cand.key), cand))
        probes.append(("tokenset", token_set(cand.key), cand))
        for short in amc_prefix_variants(cand.key, prefixes):
            if not is_specific(short):
                continue
            probes.append(("amc_prefix", short, cand))
            probes.append(("amc_prefix", despace(short), cand))

    # (rank, scoped) -> ordered list of (candidate order, candidate, scheme ids)
    buckets: dict[tuple[int, bool], dict[tuple[int, str], tuple[Candidate, set[int]]]]
    buckets = defaultdict(dict)
    for cand_tier, key, cand in probes:
        for index_tier in INDEX_TIERS:
            rank = min(TIER_RANK[cand_tier], TIER_RANK[index_tier])
            for search_scope, scoped in ((scope, True), ("", False)):
                if scoped and not scope:
                    continue
                found = universe.lookup(search_scope, index_tier, key)
                if not found:
                    continue
                slot = buckets[(rank, scoped)].setdefault(
                    (cand.order, key), (cand, set())
                )
                slot[1].update(found)

    by_name: MatchResult | None = None
    for scoped in (True, False):
        if by_name is not None:
            break
        for rank in sorted(set(TIER_RANK.values()), reverse=True):
            bucket = buckets.get((rank, scoped))
            if not bucket:
                continue
            method = RANK_METHOD[rank]
            conf = round(METHOD_CONFIDENCE[method] - (0.0 if scoped else 0.05), 2)
            # Most authoritative provenance first; the first key that resolves
            # to exactly one scheme decides the unit.
            entries = sorted(bucket.items(), key=lambda kv: kv[0])
            for (_, key), (cand, sids) in entries:
                if len(sids) != 1:
                    continue
                sid = next(iter(sids))
                evidence = f"{cand.source}: {cand.raw!r} -> {key!r}"
                if scope and universe.schemes[sid].mf_name != scope:
                    continue
                by_name = MatchResult("matched", method, conf, sid, [sid],
                                      evidence, detected=cand.raw)
                break
            if by_name is None:
                (_, key), (cand, sids) = entries[0]
                evidence = f"{cand.source}: {cand.raw!r} -> {key!r}"
                if len(sids) == 1:
                    sid = next(iter(sids))
                    by_name = MatchResult(
                        "ambiguous", method, conf, None, [sid], evidence,
                        f"name resolves to {universe.schemes[sid].mf_name or 'another AMC'},"
                        f" not {scope}",
                        detected=cand.raw,
                    )
                else:
                    by_name = MatchResult(
                        "ambiguous", method, conf, None, sorted(sids), evidence,
                        f"{len(sids)} ICRA schemes share this normalised name",
                        detected=cand.raw,
                    )
            break

    if by_name is not None and by_name.status == "matched":
        return by_name

    # Tier 4 - resolve the name through AMFI's NAVAll.txt, then match on code.
    # Worth trying even when the name lookup was ambiguous: an AMFI code is a
    # stronger signal than a normalised-name collision.
    if amfi:
        for cand in unit.candidates:
            codes = amfi.get(cand.key, set())
            if not codes:
                continue
            sids = {universe.by_code[c] for c in codes if c in universe.by_code}
            if len(sids) != 1:
                continue
            sid = next(iter(sids))
            if scope and universe.schemes[sid].mf_name != scope:
                continue
            shown_codes = ", ".join(sorted(codes)[:8])
            if len(codes) > 8:
                shown_codes += f", +{len(codes) - 8} more"
            return MatchResult(
                "matched", "amfi_navall_code", METHOD_CONFIDENCE["amfi_navall_code"],
                sid, [sid], f"{cand.raw!r} -> AMFI code(s) {shown_codes} (NAVAll.txt)",
                detected=cand.raw,
            )

    if by_name is not None:
        return by_name

    # Nothing resolved deterministically.  Surface the closest ICRA names so the
    # near miss can be adjudicated, but never accept them automatically.
    near = fuzzy_candidates(unit, universe, scope)
    detected = best_detected_name(unit, universe, scope, near[0][2] if near else "")
    if near and near[0][0] >= FUZZY_AMBIGUOUS_FLOOR:
        score, sid, raw = near[0]
        return MatchResult(
            "ambiguous", "fuzzy_near_miss", round(score, 2), None,
            [s for _, s, _ in near], f"{raw!r} (similarity {score:.2f})",
            "close but not an exact normalised match; needs manual confirmation",
            detected=detected,
            nearest_fund=universe.schemes[sid].fund_name,
            nearest_score=round(score, 2),
        )

    note = "no scheme name in this unit matched the ICRA universe"
    if not unit.candidates:
        note = "no scheme-like text found in filename, sheet name or sheet header"
    if near:
        note += (f"; closest ICRA name {universe.schemes[near[0][1]].fund_name!r}"
                 f" at {near[0][0]:.2f}")
    return MatchResult(
        "unmatched", "", 0.0, None, [], "", note, detected=detected,
        nearest_fund=universe.schemes[near[0][1]].fund_name if near else "",
        nearest_score=round(near[0][0], 2) if near else 0.0,
    )


def best_detected_name(
    unit: Unit, universe: IcraUniverse, scope: str, preferred: str = ""
) -> str:
    """Pick the most scheme-like raw string we saw for this unit.

    The text that came closest to an ICRA name wins.  Failing that, the first
    body title on the sheet is used - candidates are collected in row order, so
    the earliest one is the banner above the holdings table rather than a
    footnote further down.  Strings that are only the AMC's own name
    ("SAMCO MUTUAL FUND") identify no scheme and are skipped.
    """
    if preferred:
        return preferred
    banned = {despace(b) for b in universe.prefixes.get(scope, [])}
    banned.add(despace(canonical(scope)))

    # Among the header-area titles, the wordiest one is the scheme name; short
    # neighbours are navigation ("Back to Index") or column labels.  Footnotes
    # further down the sheet are wordier still, so they are excluded by only
    # considering the header rows that build_units already prioritised.
    best = ""
    best_tokens = 0
    fallback = ""
    for cand in unit.candidates:
        key = canonical(cand.raw)
        if not key or despace(key) in banned:
            continue
        if re.fullmatch(r"\d{4}(?:\s+\d{1,2}){2,5}", key):
            continue
        if _SECTION_NOISE.match(cand.raw.strip()):
            continue
        if not fallback:
            fallback = cand.raw
        if not cand.source.startswith(("sheet:", "index_sheet", "manifest")):
            continue
        if cand.order > ORDER_SHEET_TITLE + 20:
            continue
        tokens = len(key.split())
        if tokens > best_tokens:
            best_tokens, best = tokens, cand.raw
    return best or fallback or unit.detected_name


# ---------------------------------------------------------------------------
# Audit driver
# ---------------------------------------------------------------------------


# The two directions a coverage gap can point in.  Every report states which
# one it is, because "missing" on its own is ambiguous.
GAP_NONE = "none - matched"
GAP_REVIEW = "needs review - ambiguous"
GAP_NOT_IN_ICRA = "in data/raw, not matched to ICRA_Sample.xlsx"
GAP_NOT_DOWNLOADED = "in ICRA_Sample.xlsx, not in data/raw"
GAP_UNREADABLE = "file could not be read"

STATUS_TO_GAP = {
    "matched": GAP_NONE,
    "ambiguous": GAP_REVIEW,
    "unmatched": GAP_NOT_IN_ICRA,
    "unreadable": GAP_UNREADABLE,
}


@dataclass
class Row:
    amc_dir: str
    amc: str
    file: str
    unit: str
    detected_scheme: str
    amfi_code: str
    matched_fund: str
    method: str
    status: str
    confidence: float
    file_status: str
    gap: str
    nearest_fund: str
    nearest_score: float
    notes: str


def condense_file_results(
    results: list[tuple[Unit, MatchResult]],
) -> list[tuple[Unit, MatchResult]]:
    """Return one row per detected scheme within a downloaded file.

    Some single-scheme workbooks split equity, debt, derivatives, or issuer
    detail across multiple portfolio-style sheets.  Those are evidence for one
    scheme, not separate schemes.  Consolidated files remain one row per
    distinct matched scheme (or per distinct unresolved detected name).
    """
    groups: dict[tuple, list[tuple[Unit, MatchResult]]] = {}
    for unit, res in results:
        detected = canonical(res.detected or unit.detected_name)
        if res.status == "matched":
            key = ("matched", res.scheme_id)
        elif res.status == "ambiguous":
            key = ("ambiguous", detected or unit.label, tuple(res.candidates))
        else:
            key = ("unmatched", detected or canonical(unit.label) or unit.label)
        groups.setdefault(key, []).append((unit, res))

    out: list[tuple[Unit, MatchResult]] = []
    for members in groups.values():
        best_unit, best_res = max(
            members,
            key=lambda pair: (pair[1].confidence, bool(pair[1].evidence), -len(pair[0].label)),
        )
        labels = list(dict.fromkeys(unit.label for unit, _ in members))
        combined_unit = Unit(
            label=" | ".join(labels),
            candidates=best_unit.candidates,
            codes=sorted({code for unit, _ in members for code in unit.codes}),
            detected_name=best_unit.detected_name,
        )
        note = best_res.note
        if len(labels) > 1:
            extra = "same scheme detected in multiple workbook units: " + " | ".join(labels)
            note = "; ".join(x for x in (note, extra) if x)
        combined_res = MatchResult(
            status=best_res.status,
            method=best_res.method,
            confidence=best_res.confidence,
            scheme_id=best_res.scheme_id,
            candidates=best_res.candidates,
            evidence=best_res.evidence,
            note=note,
            detected=best_res.detected,
            nearest_fund=best_res.nearest_fund,
            nearest_score=best_res.nearest_score,
        )
        out.append((combined_unit, combined_res))
    return out


def load_manifest(period_dir: Path) -> dict[str, str]:
    """Map basename -> manifest ``scheme`` value, including archive members."""
    path = period_dir / "manifest.json"
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text())
    except Exception:
        return {}
    out: dict[str, str] = {}
    for entry in data.get("downloads", {}).values():
        scheme = entry.get("scheme") or ""
        if entry.get("path"):
            out[os.path.basename(entry["path"])] = scheme
        for member in entry.get("extracted", []) or []:
            if member.get("path"):
                out[os.path.basename(member["path"])] = scheme
    return out


def audit(args: argparse.Namespace) -> int:
    cache_dir = Path(args.cache_dir)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"Reading ICRA workbook: {args.icra}")
    portfolio, master = load_icra(Path(args.icra), args.period, cache_dir, args.refresh_cache)
    universe = build_universe(portfolio, master)
    print(f"  ICRA schemes for {args.period}: {len(universe.schemes)}"
          f" across {len(universe.mf_names)} AMCs")

    amfi = load_amfi(cache_dir, args.refresh_cache, not args.no_amfi)
    if amfi:
        print(f"  AMFI NAVAll.txt keys: {len(amfi)}")

    raw = Path(args.raw_dir)
    amc_dirs: list[str] = []
    for d in sorted(raw.iterdir(), key=lambda p: p.name):
        if not d.is_dir():
            continue
        if (d / args.period).is_dir():
            amc_dirs.append(d.name)
            continue
        # No period folder directly inside -- check one level deeper. This
        # is how Bandhan organises data/raw/bandhan/ into monthly-half-yearly/
        # and monthly/ subfolders (two structurally different kinds of
        # workbook that both belong to the same AMC), each holding its own
        # <period>/ folder. Every other AMC's period folder sits directly
        # inside its AMC folder and matches the branch above, so this is a
        # no-op for them.
        for child in sorted(d.iterdir(), key=lambda p: p.name):
            if child.is_dir() and (child / args.period).is_dir():
                amc_dirs.append(f"{d.name}/{child.name}")
    amc_dirs.sort()
    if args.amc:
        wanted = set(args.amc)
        amc_dirs = [d for d in amc_dirs if d in wanted]
    amc_map = map_amc_dirs(amc_dirs, universe.mf_names)
    for d, mf in amc_map.items():
        if not mf:
            print(f"  ! no ICRA AMC mapped for raw directory {d!r}", file=sys.stderr)

    rows: list[Row] = []
    covered: dict[int, list[str]] = defaultdict(list)
    # Schemes an ambiguous unit pointed at without being accepted - the
    # difference between "never downloaded" and "downloaded but unconfirmed".
    flagged: dict[int, list[str]] = defaultdict(list)
    per_amc: dict[str, Counter] = defaultdict(Counter)
    started = time.time()

    for amc_dir in amc_dirs:
        mf_name = amc_map.get(amc_dir, "")
        period_dir = raw / amc_dir / args.period
        manifest = load_manifest(period_dir)
        files = sorted(
            p for p in period_dir.rglob("*")
            if p.is_file() and p.name != "manifest.json" and not p.name.startswith("~$")
            and not p.name.startswith(".")
        )
        stats = per_amc[amc_dir]
        stats["files"] = len(files)
        if args.verbose:
            print(f"  [{amc_dir}] {len(files)} file(s) -> {mf_name or '<unmapped>'}")

        for path in files:
            kind, scans, error = read_workbook(path)
            rel = str(path.relative_to(raw / amc_dir / args.period))
            if error:
                stats["unreadable"] += 1
                rows.append(Row(
                    amc_dir, mf_name, rel, "", "", "", "", "", "unreadable",
                    0.0, "unreadable", GAP_UNREADABLE, "", 0.0, error,
                ))
                continue

            units, skipped = build_units(
                path, scans, manifest.get(path.name, ""), args.period
            )
            results = []
            for unit in units:
                res = match_unit(unit, universe, mf_name, amfi)
                results.append((unit, res))
            results = condense_file_results(results)

            n_matched = sum(1 for _, r in results if r.status == "matched")
            if n_matched == len(results) and n_matched:
                file_status = "fully_matched"
            elif n_matched:
                file_status = "partially_matched"
            else:
                file_status = "unmatched"
            # Prefixed so the file-level tally cannot collide with the
            # unit-level status counters ("matched"/"ambiguous"/"unmatched").
            stats["file_" + file_status] += 1

            for unit, res in results:
                scheme = universe.schemes[res.scheme_id] if res.scheme_id is not None else None
                if scheme is not None:
                    covered[scheme.scheme_id].append(f"{amc_dir}/{rel}")
                elif res.status == "ambiguous":
                    for sid in res.candidates:
                        flagged[sid].append(f"{amc_dir}/{rel}")
                notes = [res.note] if res.note else []
                if res.status == "ambiguous" and res.candidates:
                    names = [universe.schemes[i].fund_name for i in res.candidates[:5]]
                    notes.append("candidates: " + " | ".join(names))
                if res.evidence:
                    notes.append(res.evidence)
                if not scans:
                    notes.append("workbook had no sheets")
                if unit.label == "(file)" and skipped:
                    notes.append("no portfolio-style sheet; skipped: " + ", ".join(skipped[:6]))
                stats[res.status] += 1
                rows.append(
                    Row(
                        amc_dir=amc_dir,
                        amc=mf_name,
                        file=rel,
                        unit=unit.label,
                        detected_scheme=(res.detected or unit.detected_name)[:180],
                        amfi_code=scheme.amfi_code if scheme else "",
                        matched_fund=scheme.fund_name if scheme else "",
                        method=res.method,
                        status=res.status,
                        confidence=res.confidence,
                        file_status=file_status,
                        gap=STATUS_TO_GAP.get(res.status, res.status),
                        nearest_fund=res.nearest_fund,
                        nearest_score=res.nearest_score,
                        notes="; ".join(n for n in notes if n)[:600],
                    )
                )

    elapsed = time.time() - started
    print(f"  scanned {sum(s['files'] for s in per_amc.values())} files in {elapsed:.1f}s")

    write_coverage(out_dir / f"icra_coverage_{_slug(args.period)}.csv", rows)
    missing_path = out_dir / f"icra_missing_{_slug(args.period)}.csv"
    n_missing = write_missing(
        missing_path, universe, covered, flagged, amc_map, args.amc
    )
    write_unresolved(out_dir / f"icra_unresolved_downloads_{_slug(args.period)}.csv", rows)
    print_summary(rows, per_amc, universe, covered, amc_map, n_missing, args.period)
    return 0


def _slug(period: str) -> str:
    year, month = period.split("-")
    months = ["january", "february", "march", "april", "may", "june",
              "july", "august", "september", "october", "november", "december"]
    return f"{months[int(month) - 1]}_{year}"


def write_coverage(path: Path, rows: list[Row]) -> None:
    """One row per scheme detected in a downloaded file."""
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow([
            "amc", "amc_icra_name", "downloaded_file", "file_unit", "detected_scheme",
            "amfi_code", "matched_icra_fund", "match_method", "status", "confidence",
            "file_status", "gap", "closest_icra_fund", "closest_similarity", "notes",
        ])
        for r in rows:
            w.writerow([
                r.amc_dir, r.amc, r.file, r.unit, r.detected_scheme, r.amfi_code,
                r.matched_fund, r.method, r.status,
                f"{r.confidence:.2f}" if r.confidence else "", r.file_status,
                r.gap, r.nearest_fund,
                f"{r.nearest_score:.2f}" if r.nearest_score else "", r.notes,
            ])
    print(f"  wrote {path} ({len(rows)} rows)")


def write_unresolved(path: Path, rows: list[Row]) -> int:
    """Gap B: schemes found in data/raw that no ICRA scheme was matched to."""
    out = [r for r in rows if r.status in ("unmatched", "unreadable")]
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow([
            "gap", "amc", "amc_icra_name", "downloaded_file", "file_unit",
            "detected_scheme", "closest_icra_fund", "closest_similarity", "notes",
        ])
        for r in out:
            w.writerow([
                r.gap, r.amc_dir, r.amc, r.file, r.unit, r.detected_scheme,
                r.nearest_fund,
                f"{r.nearest_score:.2f}" if r.nearest_score else "", r.notes,
            ])
    print(f"  wrote {path} ({len(out)} rows)")
    return len(out)


def write_missing(
    path: Path,
    universe: IcraUniverse,
    covered: dict[int, list[str]],
    flagged: dict[int, list[str]],
    amc_map: dict[str, str],
    amc_filter: list[str] | None,
) -> int:
    """Gap A: ICRA schemes that no downloaded file was confirmed to represent."""
    audited_amcs = {mf for mf in amc_map.values() if mf}
    rows = []
    for scheme in universe.schemes:
        if scheme.scheme_id in covered:
            continue
        if amc_filter and scheme.mf_name not in audited_amcs:
            continue
        near = sorted(set(flagged.get(scheme.scheme_id, [])))
        if near:
            reason = "listed as an ambiguous candidate but not confirmed"
        elif scheme.mf_name not in audited_amcs:
            reason = "no downloaded file for this AMC in the audited period"
        else:
            reason = "AMC downloads audited but no file resolved to this scheme"
        rows.append([
            GAP_NOT_DOWNLOADED,
            scheme.mf_name or "<unknown AMC>",
            scheme.fund_name,
            scheme.amfi_code,
            reason,
            "; ".join(near[:3]),
        ])
    rows.sort(key=lambda r: (r[1], r[2]))
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow([
            "gap", "amc_icra_name", "icra_fund_name", "amfi_code", "reason",
            "ambiguous_candidate_files",
        ])
        w.writerows(rows)
    print(f"  wrote {path} ({len(rows)} rows)")
    return len(rows)


def print_summary(
    rows: list[Row],
    per_amc: dict[str, Counter],
    universe: IcraUniverse,
    covered: dict[int, list[str]],
    amc_map: dict[str, str],
    n_missing: int,
    period: str,
) -> None:
    """Print the per-AMC table, then each gap direction separately."""
    icra_by_amc = Counter(s.mf_name for s in universe.schemes)
    covered_by_amc: Counter = Counter()
    for sid in covered:
        covered_by_amc[universe.schemes[sid].mf_name] += 1

    label = period
    title = f"{label} COVERAGE: data/raw  vs  ICRA_Sample.xlsx"

    # Two blocks: what the downloaded files contain, and what ICRA expects.
    band = f"{'':<20}{'|':>1} {'DOWNLOADED (data/raw)':^70} {'|':>1} {'ICRA_Sample.xlsx':^25}"
    header = (
        f"{'AMC':<20}| {'files':>5}{'schemes':>8}{'matched':>8}{'ambig':>7}"
        f"{'unmat':>7}{'unread':>7}{'full':>6}{'part':>6}{'none':>6}"
        f" | {'total':>6}{'covered':>9}{'MISSING':>9}"
    )
    width = len(header)
    print()
    print("=" * width)
    print(title)
    print("=" * width)
    print(band)
    print(header)
    print("-" * width)

    totals = Counter()
    for amc_dir in sorted(per_amc):
        s = per_amc[amc_dir]
        mf = amc_map.get(amc_dir, "")
        found = s["matched"] + s["ambiguous"] + s["unmatched"]
        icra_n = icra_by_amc.get(mf, 0)
        cov = covered_by_amc.get(mf, 0)
        miss = icra_n - cov
        flag = " <" if miss else ""
        print(
            f"{amc_dir[:20]:<20}| {s['files']:>5}{found:>8}{s['matched']:>8}"
            f"{s['ambiguous']:>7}{s['unmatched']:>7}{s['unreadable']:>7}"
            f"{s['file_fully_matched']:>6}{s['file_partially_matched']:>6}"
            f"{s['file_unmatched']:>6} | {icra_n:>6}{cov:>9}{miss:>9}{flag}"
        )
        for k in ("files", "matched", "ambiguous", "unmatched", "unreadable",
                  "file_fully_matched", "file_partially_matched", "file_unmatched"):
            totals[k] += s[k]
        totals["found"] += found

    print("-" * width)
    print(
        f"{'TOTAL':<20}| {totals['files']:>5}{totals['found']:>8}{totals['matched']:>8}"
        f"{totals['ambiguous']:>7}{totals['unmatched']:>7}{totals['unreadable']:>7}"
        f"{totals['file_fully_matched']:>6}{totals['file_partially_matched']:>6}"
        f"{totals['file_unmatched']:>6} | "
        f"{len(universe.schemes):>6}{len(covered):>9}{n_missing:>9}"
    )
    print("=" * width)
    print(
        "files: full/part/none = every, some, or no scheme in the file matched.  "
        "schemes = distinct scheme-sized\nslices found inside each file after "
        "deduplicating repeated equity/debt/derivative tabs."
    )

    _print_gap_a(universe, covered, covered_by_amc, icra_by_amc, amc_map, n_missing, period)
    _print_gap_b(rows, period)
    _print_review(rows)

    pct = 100 * len(covered) / max(len(universe.schemes), 1)
    print()
    print("=" * width)
    print(
        f"BOTTOM LINE  {len(covered)}/{len(universe.schemes)} ICRA schemes ({pct:.1f}%) "
        f"are represented by a downloaded file."
    )
    print("=" * width)


def _print_gap_a(
    universe: IcraUniverse,
    covered: dict[int, list[str]],
    covered_by_amc: Counter,
    icra_by_amc: Counter,
    amc_map: dict[str, str],
    n_missing: int,
    period: str,
) -> None:
    """Schemes ICRA lists that nothing under data/raw was matched to."""
    print()
    print("-" * 78)
    print(f"GAP A  IN ICRA_Sample.xlsx, NOT IN data/raw   ({n_missing} schemes)")
    print("-" * 78)
    print("  ICRA lists these for the period; no downloaded file resolved to them.")
    if not n_missing:
        print("  (none)")
        return
    by_amc: dict[str, list[str]] = defaultdict(list)
    for scheme in universe.schemes:
        if scheme.scheme_id not in covered:
            by_amc[scheme.mf_name or "<unknown AMC>"].append(scheme.fund_name)
    print()
    print(f"  {'AMC':<34}{'missing':>8}{'of':>4}{'ICRA':>6}   examples")
    for mf, names in sorted(by_amc.items(), key=lambda kv: -len(kv[1])):
        eg = "; ".join(n[:34] for n in sorted(names)[:2])
        print(f"  {mf[:34]:<34}{len(names):>8}{'/':>4}{icra_by_amc.get(mf, 0):>6}   {eg[:60]}")
    print()
    print(f"  Full list: outputs/icra_missing_{_slug(period)}.csv")


def _print_gap_b(rows: list[Row], period: str) -> None:
    """Schemes present in a downloaded file that ICRA has no entry for."""
    unresolved = [r for r in rows if r.status in ("unmatched", "unreadable")]
    print()
    print("-" * 78)
    print(f"GAP B  IN data/raw, NOT MATCHED TO ICRA_Sample.xlsx   ({len(unresolved)} slices)")
    print("-" * 78)
    print("  Found inside a downloaded file, but no ICRA scheme was confirmed for it.")
    if not unresolved:
        print("  (none)")
        return
    print()
    for r in unresolved:
        name = r.detected_scheme or r.unit or "(no scheme name found)"
        print(f"  [{r.amc_dir}] {name[:62]}")
        if r.nearest_fund:
            print(f"      closest ICRA name: {r.nearest_fund[:52]!r} ({r.nearest_score:.2f})")
        else:
            print("      no similar ICRA name - likely absent from ICRA entirely")
        print(f"      file: {r.file[:62]}")
    print()
    print(f"  Full list: outputs/icra_unresolved_downloads_{_slug(period)}.csv")


def _print_review(rows: list[Row]) -> None:
    """Near misses deliberately not auto-accepted."""
    amb = [r for r in rows if r.status == "ambiguous"]
    print()
    print("-" * 78)
    print(f"NEEDS REVIEW  close but not accepted automatically   ({len(amb)} slices)")
    print("-" * 78)
    if not amb:
        print("  (none)")
        return
    print("  Counted as a gap in BOTH directions until a human confirms them.")
    print()
    for r in amb:
        print(f"  [{r.amc_dir}] {r.detected_scheme[:58]}")
        print(f"      probably: {r.nearest_fund[:56]!r} ({r.confidence:.2f})")


def main(argv: Sequence[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--period", default=DEFAULT_PERIOD, help="YYYY-MM (default %(default)s)")
    p.add_argument("--icra", default=str(DEFAULT_ICRA))
    p.add_argument("--raw-dir", default=str(DEFAULT_RAW))
    p.add_argument("--out-dir", default=str(DEFAULT_OUT))
    p.add_argument("--cache-dir", default=str(DEFAULT_CACHE))
    p.add_argument("--amc", action="append", help="restrict to one or more raw AMC directories")
    p.add_argument("--refresh-cache", action="store_true", help="re-read ICRA and re-fetch AMFI")
    p.add_argument("--no-amfi", action="store_true", help="skip the AMFI NAVAll.txt fallback")
    p.add_argument("--verbose", action="store_true")
    args = p.parse_args(argv)
    return audit(args)


if __name__ == "__main__":
    raise SystemExit(main())
