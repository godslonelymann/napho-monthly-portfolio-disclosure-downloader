"""360 ONE Mutual Fund — monthly portfolio parser.

Which files/sheets are schemes: the raw/360_one/<period>/ folder holds a lot
of unrelated downloads (daily AAUM disclosures, the AMC dashboard, portfolio
overlap reports, single-scheme YTM sheets). The one true monthly portfolio
file is named "..._MONTHLY_PORTFOLIO_<Mon><YYYY>..._Final....xlsx" and has
one sheet per scheme — no index sheet, no multi-scheme sheets.

Header row: row 4 in every sheet ("Name of the Instrument", "ISIN", ...).
Row 1 is the scheme's full title, row 3 is "Monthly Portfolio Statement as
on <date>". Column layout is fixed across all 12 sheets:

    A: internal broker/security code (unused)
    B: Name of the Instrument
    C: ISIN
    D: Industry / Rating (raw, AMC's own wording)
    E: Quantity
    F: Market/Fair Value (Rs. in Lacs)
    G: Rounded % to Net Assets (fraction, e.g. 0.0525 = 5.25%)
    H: YTM (unused)
    I: ~YTC (unused)

Section headers: rows where column B has text but E/F/G are all empty.
Top-level headers ("Equity & Equity related", "Debt Instruments", "Money
Market Instruments", "TREPS / Reverse Repo", "Gold", "REIT/InvIT
Instruments", "Others") and finer sub-labels ("Certificate of Deposit",
"Commercial Paper", "Treasury Bill", ...) are both tracked into
section_header for context/debugging, but pipeline/convert.py does NOT rely
on section_header to determine Instrument_Name — ISIN determines it
deterministically dataset-wide (verified against ICRA_Sample.xlsx: 0 ISINs
in 128k rows map to more than one Instrument_Name), and the handful of
non-ISIN rows (TREPS, Reverse Repo, Gold, Silver, Net Receivables,
commodity futures) are keyed off their own security_name instead — see
data/mappings/360_one/non_isin_instruments.csv. One AMC quirk this
sidesteps: 360 ONE's own file mislabels the Silver ETF's section header as
"Gold" (template copy-paste), which section-header-based classification
would get wrong.

"(a) Listed / awaiting listing..." / "(b) Privately placed / Unlisted" rows
carry no type information (just exchange-listing status) and are skipped.
Rows literally named "Sub Total" / "Total" / "GRAND TOTAL" are skipped.
Parsing of a sheet stops at "Notes:" — everything after is NAV/IDCW/YTM
footnotes, not holdings.

Units (handled in pipeline/convert.py, not here): column F is Rs. in Lacs
(-> divide by 100 for the crores ICRA expects); column G is a fraction
(0.0525, not 5.25) so it's multiplied by 100 for Corpus_Per. This parser
hands both through unconverted — converting here would make the
intermediate CSV lie about what's actually in the source file.

ISIN drift vs. ICRA_Sample.xlsx: two securities in this file (Talwandi
Sabo Power, Info Edge) carry an ISIN that differs from the one ICRA's May
2026 sample uses for the same holding — same quantity and value on both
sides, just a different ISIN string (Vedanta demerger fallout in one
case). This is not a parsing bug; see data/lookups/isin_aliases.csv and
pipeline/convert.py, which resolve it at the shared-converter layer, not
per-AMC.

What every future AMC parser should copy from this one: find the header
row by label, not by row number; keep classification keyed off ISIN, not
section labels — 360 ONE's own file mislabels the Silver ETF's section as
"Gold" (template copy-paste), and header-based logic would get it wrong.
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from pipeline.schema import IntermediateRow

AMC = "360_one"

_MONTHLY_FILE_RE = re.compile(r"MONTHLY_PORTFOLIO", re.IGNORECASE)

TOP_LEVEL_SECTIONS = {
    "equity & equity related",
    "debt instruments",
    "money market instruments",
    "treps / reverse repo",
    "gold",
    "reit/invit instruments",
    "others",
}

SUBSECTION_LABELS = {
    "certificate of deposit",
    "commercial paper",
    "treasury bill",
    "corporate debt market development fund",
    "exchange traded funds",
    "commodity future",
}

TOTAL_WORDS = {"sub total", "total", "grand total"}
STOP_MARKERS = {"notes:"}
SKIP_LABEL_PREFIXES = ("(a)", "(b)")


def find_monthly_file(period_dir: str | Path) -> Path:
    period_dir = Path(period_dir)
    candidates = [
        p for p in period_dir.glob("*.xlsx") if _MONTHLY_FILE_RE.search(p.name)
    ]
    if not candidates:
        raise FileNotFoundError(f"No monthly-portfolio workbook found in {period_dir}")
    if len(candidates) > 1:
        raise RuntimeError(f"Ambiguous monthly-portfolio workbooks in {period_dir}: {candidates}")
    return candidates[0]


def _find_header_row(ws) -> int:
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == "Name of the Instrument":
                return cell.row
    raise ValueError(f"Header row not found in sheet {ws.title!r}")


def _parse_sheet(ws, *, amc: str, source_file: str) -> list[IntermediateRow]:
    header_row = _find_header_row(ws)
    scheme_name_raw = ws.title
    section: str | None = None
    subsection: str | None = None
    out: list[IntermediateRow] = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = row[1] if len(row) > 1 else None
        if name is None:
            continue
        key = str(name).strip()
        if not key:
            continue
        key_norm = key.lower()

        if key_norm in TOTAL_WORDS:
            continue
        if key.startswith(SKIP_LABEL_PREFIXES):
            continue
        if key_norm in STOP_MARKERS:
            break

        isin = row[2] if len(row) > 2 else None
        industry_raw = row[3] if len(row) > 3 else None
        quantity = row[4] if len(row) > 4 else None
        market_value_raw = row[5] if len(row) > 5 else None
        pct_raw = row[6] if len(row) > 6 else None
        has_value = any(v is not None for v in (quantity, market_value_raw, pct_raw))

        if not has_value:
            if key_norm in TOP_LEVEL_SECTIONS:
                section = key
                subsection = None
            elif key_norm in SUBSECTION_LABELS:
                subsection = key
            # else: unrecognized footnote/label row — ignore
            continue

        out.append(
            IntermediateRow(
                amc=amc,
                source_file=source_file,
                sheet=ws.title,
                scheme_name_raw=scheme_name_raw,
                section_header=subsection or section,
                security_name=key,
                isin=(str(isin).strip() if isin else None),
                industry_raw=(str(industry_raw).strip() if industry_raw not in (None, "") else None),
                quantity=quantity,
                market_value_raw=market_value_raw,
                pct_raw=pct_raw,
            )
        )

    return out


def parse_workbook(path: str | Path) -> list[IntermediateRow]:
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    out: list[IntermediateRow] = []
    for sheet_name in wb.sheetnames:
        out.extend(_parse_sheet(wb[sheet_name], amc=AMC, source_file=path.name))
    return out


def parse_period(period_dir: str | Path) -> list[IntermediateRow]:
    return parse_workbook(find_monthly_file(period_dir))


if __name__ == "__main__":
    import sys

    period_dir = sys.argv[1] if len(sys.argv) > 1 else "data/raw/360_one/2026-05"
    rows = parse_period(period_dir)
    print(f"{len(rows)} rows from {find_monthly_file(period_dir).name}")
    schemes = sorted({r.sheet for r in rows})
    print(f"{len(schemes)} schemes: {schemes}")
