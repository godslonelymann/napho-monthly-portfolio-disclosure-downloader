"""Build the ISIN -> security-name table the ICRA sheet doesn't give us.

ICRA_Sample.xlsx's portfolio sheet identifies a holding only by ISIN;
there is no column anywhere in it for "Vardhman Textiles Limited". The
AMC files, on the other hand, print the name on every holding row right
next to the ISIN. So the mapping is already sitting in data/raw — this
module harvests it rather than fetching it from anywhere.

Two outputs, because the mapping is needed in both directions:

  isin_names.csv     ISIN -> one canonical name (plus how many times it
                     was seen and how many spellings exist). This is the
                     enrichment direction.

  isin_name_variants.csv
                     every raw spelling ever observed, with counts. Not
                     noise — it is the *matching* direction. 10.4% of
                     ICRA's rows (13,309 of 127,912: Futures, Money
                     Market, Net Receivables, IRS, Margin Deposit, Cash,
                     Repo) carry no ISIN at all, so a name is the only
                     handle we have on them, and AMCs spell the same
                     holding up to 20 different ways.

  unlinked_names.csv names seen on a holding row that had no ISIN, with
                     counts and the section header they sat under. The
                     raw material for resolving that 10.4%.

Canonical-name choice: variants are grouped by a match_key (uppercased,
punctuation and footnote marks stripped, Ltd/Ltd./Limited folded
together) so that "HDFC Bank Ltd." and "HDFC BANK LIMITED" land in one
bucket. Within a bucket the most frequently observed raw spelling wins,
ties broken towards the longer form ("Limited" over "Ltd"). Corporate
renames are deliberately NOT collapsed — Eternal Limited still shows its
old "Zomato Ltd" spellings, which is what lets an older file match.
"""

from __future__ import annotations

import csv
import re
import sys
import warnings
from collections import Counter, defaultdict
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

warnings.filterwarnings("ignore")

# --- what a holding row looks like ------------------------------------

# ISO 6166 shape: 2-letter country code + 9 alphanumeric + 1 check digit.
# Deliberately NOT restricted to ^IN — international and overseas funds
# hold US, IE, LU and other foreign lines (Microsoft US5949181045,
# Cognizant US1924461023), and an India-only pattern silently drops them.
#
# Shape alone is NOT enough. Derivative contract codes are also 12
# characters and match it: AUBANK300626 (AU Bank future, 30-06-26),
# BHEL29052025, SBIN26122024, plus template leftovers like ABCDE1234567.
# 1,200 of these got into an earlier build of the table, some carrying
# junk names lifted from the Long/Short column. The check digit is what
# separates them, so ISIN_SHAPE is only ever used via is_isin().
ISIN_SHAPE = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}[0-9]$")


def isin_check_digit_ok(code: str) -> bool:
    """ISO 6166 check digit: letters -> two digits (A=10..Z=35), then Luhn."""
    digits = ""
    for ch in code:
        if ch.isdigit():
            digits += ch
        elif ch.isalpha():
            digits += str(ord(ch) - 55)
        else:
            return False
    total = 0
    for i, ch in enumerate(reversed(digits)):
        d = int(ch)
        if i % 2 == 1:
            d *= 2
            if d > 9:
                d -= 9
        total += d
    return total % 10 == 0


def is_isin(value: str) -> bool:
    return bool(ISIN_SHAPE.match(value)) and isin_check_digit_ok(value)

# The 136 distinct header labels across all 52 AMCs collapse onto these.
NAME_HDR = re.compile(
    r"name\s*of\s*(the\s*)?instrument"
    r"|name\s*of\s*(the\s*)?instrument\s*/\s*issuer"
    r"|instrument\s*name"
    r"|company\s*/\s*issuer"
    r"|name\s*of\s*(the\s*)?security"
    r"|particulars",
    re.I,
)
ISIN_HDR = re.compile(r"^isin(\s*(code|number|no\.?))?$", re.I)

# Rows that are structure, not holdings.
SKIP_ROW = re.compile(
    r"^\s*(sub\s*total|total|grand\s*total|net\s*assets?"
    r"|\(?[a-d]\)?\s*(listed|unlisted|privately|foreign)"
    r"|notes?\s*:|nil)\s*$",
    re.I,
)

# Trailing/leading junk AMCs append to names: footnote daggers, exchange
# prefixes, and the expiry stamp that marks a derivative line.
FOOTNOTE_CHARS = "*^£$#~!@%&+°·•"
DERIV_SUFFIX = re.compile(r"\s*[-–]\s*\d{1,2}[-/][A-Za-z]{3}[-/]\d{2,4}\s*$")
EQ_PREFIX = re.compile(r"^\s*(EQ|FUT|OPT|CASH)\s*[-–]\s*", re.I)
CORP_SUFFIX = re.compile(r"\b(LIMITED|LTD)\b\.?", re.I)


def clean_name(raw: str) -> tuple[str, bool]:
    """Strip decoration off a raw name. Returns (name, is_derivative)."""
    name = " ".join(str(raw).split())
    name = EQ_PREFIX.sub("", name)
    is_deriv = bool(DERIV_SUFFIX.search(name))
    name = DERIV_SUFFIX.sub("", name)
    # Footnote marks cluster at the end, sometimes several with spaces
    # between them ("Reliance Industries Ltd $$ ~~").
    while True:
        stripped = name.rstrip(FOOTNOTE_CHARS + " ")
        if stripped == name:
            break
        name = stripped
    return name.strip(), is_deriv


def match_key(name: str) -> str:
    """Fold spellings of the same entity together for grouping only."""
    key = CORP_SUFFIX.sub("LIMITED", name.upper())
    key = re.sub(r"[^A-Z0-9]+", "", key)
    return key


# --- reading a workbook of either format ------------------------------


def _iter_sheets_xlsx(path: Path):
    import io

    import openpyxl

    # Handed the raw bytes rather than the path on purpose: openpyxl
    # rejects anything named *.xls by extension alone, and half the .xls
    # files here are really xlsx (see iter_sheets).
    wb = openpyxl.load_workbook(
        io.BytesIO(path.read_bytes()), read_only=True, data_only=True
    )
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            yield sheet_name, (list(r) for r in ws.iter_rows(values_only=True))
    finally:
        wb.close()


def _iter_sheets_xls(path: Path):
    import xlrd

    wb = xlrd.open_workbook(path, on_demand=True)
    try:
        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            yield sheet_name, (ws.row_values(i) for i in range(ws.nrows))
    finally:
        wb.release_resources()


def iter_sheets(path: Path):
    """Dispatch on the file's actual format, not its extension.

    Several AMCs ship a real .xlsx under a .xls name (nippon_india,
    baroda_bnp, ppfas, abakkus all do in May 2026), so trusting the
    suffix loses four of the larger AMCs outright. Magic bytes:
    PK\\x03\\x04 = zip = xlsx; \\xd0\\xcf\\x11\\xe0 = OLE2 = real xls.
    """
    with path.open("rb") as f:
        magic = f.read(4)
    if magic == b"\xd0\xcf\x11\xe0":
        return _iter_sheets_xls(path)
    return _iter_sheets_xlsx(path)


# --- the harvest ------------------------------------------------------


def _locate_columns(rows: list[list]) -> tuple[int, int] | None:
    """Find (name_col, isin_col) for a sheet by looking at the data.

    Header position is not trustworthy. Kotak's header puts "Name of
    Instrument" in column 0 while every actual name sits in column 2;
    several AMCs carry an unlabelled internal security code that shifts
    data one column right of its own header. So the ISIN column is found
    by regex over the cells, and the name column is then chosen from the
    remaining columns by *uniqueness*: names are near-unique per row,
    whereas the industry/rating column beside them repeats ("Banks"
    hundreds of times).
    """
    isin_hits: Counter = Counter()
    for row in rows:
        for idx, cell in enumerate(row):
            if isinstance(cell, str) and is_isin(cell.strip().upper()):
                isin_hits[idx] += 1
    if not isin_hits:
        return None
    isin_col = isin_hits.most_common(1)[0][0]

    # Only judge columns on rows that actually carry an ISIN.
    data_rows = [
        row
        for row in rows
        if isin_col < len(row)
        and isinstance(row[isin_col], str)
        and is_isin(row[isin_col].strip().upper())
    ]
    if not data_rows:
        return None

    best_col, best_score = None, 0.0
    width = max(len(row) for row in data_rows)
    for idx in range(width):
        if idx == isin_col:
            continue
        values = [
            " ".join(row[idx].split())
            for row in data_rows
            if idx < len(row) and isinstance(row[idx], str) and row[idx].strip()
        ]
        values = [v for v in values if len(v) >= 4 and not ISIN_SHAPE.match(v.upper())]
        if len(values) < max(1, len(data_rows) * 0.5):
            continue
        uniqueness = len(set(values)) / len(values)
        avg_len = sum(len(v) for v in values) / len(values)
        # Uniqueness dominates; length breaks ties between a name column
        # and a sparse code column. Proximity to the ISIN nudges the
        # winner, since name and ISIN are almost always adjacent.
        score = uniqueness * min(avg_len, 40) / (1 + 0.15 * abs(idx - isin_col))
        if score > best_score:
            best_col, best_score = idx, score
    if best_col is None:
        return None
    return best_col, isin_col


def harvest_file(path: Path, amc: str, period: str) -> dict:
    """Pull every (isin, name) and every unlinked name out of one file."""
    linked: Counter = Counter()  # (isin, raw_name, is_deriv) -> n
    contracts: Counter = Counter()  # (contract_code, raw_name) -> n
    unlinked: Counter = Counter()  # (raw_name, section) -> n

    for _sheet_name, row_iter in iter_sheets(path):
        rows = [list(r) for r in row_iter]
        located = _locate_columns(rows)
        if located is None:
            continue
        name_col, isin_col = located

        section = ""
        for row in rows:
            if name_col >= len(row):
                continue
            raw_name = row[name_col]
            if not isinstance(raw_name, str) or not raw_name.strip():
                continue
            if SKIP_ROW.match(raw_name):
                continue
            if NAME_HDR.search(raw_name) or ISIN_HDR.match(raw_name.strip()):
                continue

            isin = row[isin_col] if isin_col < len(row) else None
            isin = isin.strip().upper() if isinstance(isin, str) else ""

            name, is_deriv = clean_name(raw_name)
            if not name or len(name) < 3:
                continue

            if is_isin(isin):
                linked[(isin, name, is_deriv)] += 1
            elif ISIN_SHAPE.match(isin):
                # Right shape, wrong check digit: a derivative contract
                # code, not an ISIN. Kept separately - these are the
                # handle on the Futures rows ICRA carries without an ISIN.
                contracts[(isin, name)] += 1
            else:
                # No ISIN. Either a section header or one of the ~10% of
                # holdings ICRA also carries without an ISIN.
                numeric = sum(
                    1 for c in row if isinstance(c, (int, float)) and c not in (0, 1)
                )
                if numeric >= 2:
                    unlinked[(name, section)] += 1
                else:
                    section = name

    return {
        "amc": amc,
        "period": period,
        "linked": linked,
        "contracts": contracts,
        "unlinked": unlinked,
        "path": str(path),
    }


def _worker(args) -> dict | None:
    path, amc, period = args
    try:
        return harvest_file(Path(path), amc, period)
    except Exception as exc:  # a corrupt/locked workbook must not kill the run
        return {"amc": amc, "period": period, "path": str(path), "error": repr(exc)}


def find_inputs(raw_dir: Path, periods: list[str] | None = None):
    for amc_dir in sorted(raw_dir.iterdir()):
        if not amc_dir.is_dir():
            continue
        for period_dir in sorted(amc_dir.iterdir()):
            if not period_dir.is_dir():
                continue
            if periods and period_dir.name not in periods:
                continue
            for path in sorted(period_dir.iterdir()):
                if path.suffix.lower() in (".xlsx", ".xls"):
                    yield (str(path), amc_dir.name, period_dir.name)


def build(
    raw_dir: str | Path = "data/raw",
    out_dir: str | Path = "data/lookups",
    periods: list[str] | None = None,
    workers: int = 8,
) -> dict:
    raw_dir, out_dir = Path(raw_dir), Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    inputs = list(find_inputs(raw_dir, periods))
    print(f"harvesting {len(inputs)} workbooks with {workers} workers", flush=True)

    # isin -> raw_name -> count ; plus provenance
    variants: dict[str, Counter] = defaultdict(Counter)
    isin_amcs: dict[str, set] = defaultdict(set)
    isin_periods: dict[str, set] = defaultdict(set)
    # isin -> name -> most recent period it was seen spelled that way.
    # Corporate renames (Zomato Ltd -> Eternal Limited) mean the
    # most-frequent spelling across all history is not the same thing as
    # the *current* one — an old file can out-vote a rename that only
    # covers a few recent months. Recency, not frequency, decides the
    # canonical spelling; frequency only breaks ties within one period.
    name_last_period: dict[str, dict[str, str]] = defaultdict(dict)
    deriv_names: Counter = Counter()
    contracts: dict[str, Counter] = defaultdict(Counter)
    unlinked: Counter = defaultdict(Counter)
    errors: list[tuple[str, str]] = []
    done = 0

    with ProcessPoolExecutor(max_workers=workers) as pool:
        futures = [pool.submit(_worker, a) for a in inputs]
        for fut in as_completed(futures):
            res = fut.result()
            done += 1
            if done % 2000 == 0:
                print(f"  {done}/{len(inputs)} files, {len(variants)} ISINs", flush=True)
            if res is None:
                continue
            if "error" in res:
                errors.append((res["path"], res["error"]))
                continue
            for (isin, name, is_deriv), n in res["linked"].items():
                variants[isin][name] += n
                isin_amcs[isin].add(res["amc"])
                isin_periods[isin].add(res["period"])
                prev = name_last_period[isin].get(name)
                if prev is None or res["period"] > prev:
                    name_last_period[isin][name] = res["period"]
                if is_deriv:
                    deriv_names[(isin, name)] += n
            for (code, name), n in res.get("contracts", {}).items():
                contracts[code][name] += n
            for (name, section), n in res["unlinked"].items():
                unlinked[name][section] += n

    print(f"  {done}/{len(inputs)} files, {len(variants)} ISINs, {len(errors)} errors", flush=True)

    # --- pick a canonical name per ISIN --------------------------------
    canon_rows = []
    variant_rows = []
    for isin in sorted(variants):
        counter = variants[isin]
        groups: dict[str, Counter] = defaultdict(Counter)
        for name, n in counter.items():
            groups[match_key(name)][name] += n
        # Winning entity = whichever match_key was seen most recently
        # (ties broken by total observations) — an entity that renamed
        # still wins on its new spelling's recency even if the old
        # spelling has more historical rows.
        best_key = max(
            groups,
            key=lambda k: (
                max(name_last_period[isin].get(name, "") for name in groups[k]),
                sum(groups[k].values()),
            ),
        )
        best = groups[best_key]
        # Within it, most-recently-seen spelling; ties go to more
        # observations, then the longer form.
        canonical = max(
            best.items(),
            key=lambda kv: (name_last_period[isin].get(kv[0], ""), kv[1], len(kv[0])),
        )[0]
        periods_seen = sorted(isin_periods[isin])
        canon_rows.append(
            {
                "ISIN": isin,
                "canonical_name": canonical,
                "observations": sum(counter.values()),
                "variants": len(counter),
                "entities": len(groups),
                "amcs": len(isin_amcs[isin]),
                "first_period": periods_seen[0] if periods_seen else "",
                "last_period": periods_seen[-1] if periods_seen else "",
            }
        )
        for name, n in counter.most_common():
            variant_rows.append(
                {
                    "ISIN": isin,
                    "name": name,
                    "match_key": match_key(name),
                    "count": n,
                    "is_canonical": int(name == canonical),
                }
            )

    _write(out_dir / "isin_names.csv", canon_rows)
    _write(out_dir / "isin_name_variants.csv", variant_rows)
    _write(
        out_dir / "contract_codes.csv",
        [
            {
                "contract_code": code,
                "underlying_name": names.most_common(1)[0][0],
                "observations": sum(names.values()),
                "variants": len(names),
            }
            for code, names in sorted(
                contracts.items(), key=lambda kv: -sum(kv[1].values())
            )
        ],
    )
    _write(
        out_dir / "unlinked_names.csv",
        [
            {
                "name": name,
                "count": sum(sections.values()),
                "top_section": sections.most_common(1)[0][0] if sections else "",
            }
            for name, sections in sorted(
                unlinked.items(), key=lambda kv: -sum(kv[1].values())
            )
        ],
    )
    if errors:
        _write(
            out_dir / "isin_names_errors.csv",
            [{"path": p, "error": e} for p, e in errors],
        )

    stats = {
        "files": len(inputs),
        "files_failed": len(errors),
        "isins": len(canon_rows),
        "variant_rows": len(variant_rows),
        "unlinked_names": len(unlinked),
        "contract_codes": len(contracts),
        "multi_entity_isins": sum(1 for r in canon_rows if r["entities"] > 1),
    }
    return stats


def _write(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    with path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


if __name__ == "__main__":
    periods = sys.argv[1:] or None
    stats = build(periods=periods)
    for key, value in stats.items():
        print(f"{key}: {value}")
